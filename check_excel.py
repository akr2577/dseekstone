import openpyxl
import json

# Load the Excel file
wb = openpyxl.load_workbook('stones_data_hard_pow.xlsx')
ws = wb.active

# Get headers
headers = []
for cell in ws[1]:
    headers.append(cell.value)

print("Column headers:")
for i, header in enumerate(headers, 1):
    print(f"{i}: {header}")

print("\n\nFirst 10 rows of data:")
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=11, values_only=True), 2):
    print(f"\nRow {row_idx}:")
    for i, value in enumerate(row):
        if value is not None:
            print(f"  {headers[i]}: {value}")

# Count rows with missing shardness or spower
shardness_col = headers.index('shardness') if 'shardness' in headers else None
spower_col = headers.index('spower') if 'spower' in headers else None

print("\n\nChecking for missing shardness/spower values:")
missing_count = 0
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
    if row[0] is None:  # End of data
        break
    shardness_missing = shardness_col is not None and (row[shardness_col] is None or row[shardness_col] == '')
    spower_missing = spower_col is not None and (row[spower_col] is None or row[spower_col] == '')
    
    if shardness_missing or spower_missing:
        missing_count += 1
        stone_name = row[0] if row[0] else "Unknown"
        print(f"Row {row_idx + 1}: {stone_name} - shardness: {row[shardness_col] if shardness_col else 'N/A'}, spower: {row[spower_col] if spower_col else 'N/A'}")

print(f"\nTotal rows with missing values: {missing_count}")
