import openpyxl
from openpyxl.utils import get_column_letter
import json

# โหลดไฟล์ Excel
wb = openpyxl.load_workbook('stones_data.xlsx')
ws = wb.active

# ดึง headers
headers = [cell.value for cell in ws[1]]

print("=" * 120)
print("COLUMN ANALYSIS:")
print("=" * 120)
for idx, header in enumerate(headers, 1):
    print(f"{idx:3d}. {header}")

# ค้นหาคอลัมน์ที่ต้องการ
stone_name_col = None
shardness_col = None
spower_col = None

for idx, header in enumerate(headers, 1):
    if header:
        h_lower = str(header).lower()
        if 'english_name' in h_lower or 'name' in h_lower:
            stone_name_col = idx
        if 'shardness' in h_lower or 'hardness' in h_lower:
            shardness_col = idx
        if 'spower' in h_lower or 'power' in h_lower:
            spower_col = idx

print("\n" + "=" * 120)
print(f"Stone Name Column: {stone_name_col}")
print(f"Shardness Column: {shardness_col}")
print(f"Spower Column: {spower_col}")

# ตรวจสอบค่าที่ขาดหายไป
print("\n" + "=" * 120)
print("ROWS WITH MISSING OR INCORRECT VALUES:")
print("=" * 120)

missing_data = []
for row_idx in range(2, ws.max_row + 1):
    stone_name = ws.cell(row_idx, stone_name_col).value if stone_name_col else None
    shardness = ws.cell(row_idx, shardness_col).value if shardness_col else None
    spower = ws.cell(row_idx, spower_col).value if spower_col else None
    
    if not stone_name:
        break
    
    # เก็บข้อมูลที่ขาดหายหรือค่า 0
    if (shardness is None or shardness == '' or shardness == 0 or 
        spower is None or spower == '' or spower == 0):
        missing_data.append({
            'row': row_idx,
            'stone_name': stone_name,
            'shardness': shardness,
            'spower': spower
        })
        print(f"Row {row_idx:3d}: {stone_name:40s} | Hardness: {str(shardness):5s} | Power: {str(spower):5s}")

print(f"\nTotal missing/empty: {len(missing_data)} rows")

# แสดง 20 แถวแรก
print("\n" + "=" * 120)
print("FIRST 20 ROWS:")
print("=" * 120)
for row_idx in range(2, min(22, ws.max_row + 1)):
    stone_name = ws.cell(row_idx, stone_name_col).value if stone_name_col else None
    shardness = ws.cell(row_idx, shardness_col).value if shardness_col else None
    spower = ws.cell(row_idx, spower_col).value if spower_col else None
    print(f"Row {row_idx:3d}: {stone_name:40s} | Hardness: {str(shardness):5s} | Power: {str(spower):5s}")
