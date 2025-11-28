import openpyxl
import json

# โหลดไฟล์ Excel
wb = openpyxl.load_workbook('stones_data_hard_pow.xlsx')
ws = wb.active

# ดึงข้อมูลแถวแรก (headers)
headers = [cell.value for cell in ws[1]]
print("=" * 100)
print("COLUMN HEADERS:")
for idx, header in enumerate(headers, 1):
    print(f"{idx:3d}. {header}")

# ดึงข้อมูล 10 แถวแรก
print("\n" + "=" * 100)
print("FIRST 10 ROWS OF DATA:")
for row_idx in range(2, 12):
    row_data = [cell.value for cell in ws[row_idx]]
    print(f"\nRow {row_idx}: {row_data[0] if row_data else 'EMPTY'}")
    if len(row_data) > 10:
        print(f"  Columns 1-10: {row_data[0:10]}")
        print(f"  Columns 11+: {row_data[10:20]}")

# ค้นหา columns ที่เกี่ยวข้อง
print("\n" + "=" * 100)
hardness_col = None
power_col = None
stone_name_col = None

for idx, header in enumerate(headers):
    if header and isinstance(header, str):
        if 'hardness' in header.lower() or 'shardness' in header.lower():
            hardness_col = idx
            print(f"Found HARDNESS column at index {idx}: {header}")
        if 'power' in header.lower() or 'spower' in header.lower():
            power_col = idx
            print(f"Found POWER column at index {idx}: {header}")
        if 'name' in header.lower() or 'stone' in header.lower():
            stone_name_col = idx

# ตรวจสอบค่าที่ขาดหายไป
print("\n" + "=" * 100)
print("MISSING VALUES ANALYSIS:")
missing_hardness = 0
missing_power = 0
total_rows = 0

for row_idx in range(2, ws.max_row + 1):
    total_rows += 1
    hardness_val = ws.cell(row_idx, hardness_col + 1).value if hardness_col is not None else None
    power_val = ws.cell(row_idx, power_col + 1).value if power_col is not None else None
    stone_name = ws.cell(row_idx, 1).value if stone_name_col is not None else None
    
    if hardness_val is None or hardness_val == '':
        missing_hardness += 1
        if missing_hardness <= 5:
            print(f"Row {row_idx}: {stone_name} - Missing HARDNESS")
    
    if power_val is None or power_val == '':
        missing_power += 1
        if missing_power <= 5:
            print(f"Row {row_idx}: {stone_name} - Missing POWER")

print(f"\nTotal rows: {total_rows}")
print(f"Missing HARDNESS: {missing_hardness}")
print(f"Missing POWER: {missing_power}")
